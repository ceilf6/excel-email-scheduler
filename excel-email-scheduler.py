#!/usr/bin/env python3
"""
Excel Email Scheduler

This script reads all xlsx files in the workspace directory and sends emails
based on a configurable template. Email addresses are extracted from columns
containing 'Leader' in their names.

Supports multiple authentication methods:
- Basic SMTP (username/password) for QQ Mail, 163 Mail, etc.
- Gmail OAuth2 for Gmail accounts
- Outlook OAuth2 for Outlook/Office365 accounts

Usage:
    python excel-email-scheduler.py [--config CONFIG_PATH] [--dry-run]
"""

import argparse
import base64
import json
import logging
import os
import re
import signal
import smtplib
import sys
import time
from abc import ABC, abstractmethod
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook


# =============================================================================
# Email Sender Abstract Base Class
# =============================================================================

class EmailSender(ABC):
    """Abstract base class for email senders."""

    @abstractmethod
    def send(
        self,
        from_addr: str,
        to_addr: str,
        subject: str,
        body: str,
        logger: logging.Logger
    ) -> bool:
        """Send an email."""
        pass

    @abstractmethod
    def close(self) -> None:
        """Clean up resources."""
        pass


# =============================================================================
# Basic SMTP Email Sender
# =============================================================================

class BasicSMTPSender(EmailSender):
    """Email sender using basic SMTP authentication."""

    def __init__(self, config: dict):
        """
        Initialize SMTP sender.

        Args:
            config: SMTP configuration dictionary
        """
        self.host = config["host"]
        self.port = config["port"]
        self.use_tls = config.get("use_tls", True)
        self.username = config["username"]
        self.password = config["password"]

    def send(
        self,
        from_addr: str,
        to_addr: str,
        subject: str,
        body: str,
        logger: logging.Logger
    ) -> bool:
        """Send email using SMTP."""
        try:
            msg = MIMEMultipart()
            msg["From"] = from_addr
            msg["To"] = to_addr
            msg["Subject"] = subject
            msg.attach(MIMEText(body, "plain", "utf-8"))

            if self.use_tls:
                server = smtplib.SMTP(self.host, self.port)
                server.starttls()
            else:
                server = smtplib.SMTP(self.host, self.port)

            server.login(self.username, self.password)
            server.sendmail(from_addr, to_addr, msg.as_string())
            server.quit()

            logger.info(f"Email sent successfully to: {to_addr}")
            return True

        except smtplib.SMTPException as e:
            logger.error(f"Failed to send email to {to_addr}: {e}")
            return False

    def close(self) -> None:
        """No cleanup needed for basic SMTP."""
        pass


# =============================================================================
# Gmail OAuth2 Email Sender
# =============================================================================

class GmailOAuth2Sender(EmailSender):
    """Email sender using Gmail OAuth2 authentication."""

    def __init__(self, config: dict, logger: logging.Logger):
        """
        Initialize Gmail OAuth2 sender.

        Args:
            config: Gmail OAuth2 configuration dictionary
            logger: Logger instance
        """
        try:
            from google.auth.transport.requests import Request
            from google.oauth2.credentials import Credentials
            from google_auth_oauthlib.flow import InstalledAppFlow
        except ImportError:
            raise ImportError(
                "Gmail OAuth2 requires google-auth-oauthlib. "
                "Install with: pip install google-auth-oauthlib google-auth"
            )

        self.credentials_file = config["credentials_file"]
        self.token_file = config["token_file"]
        self.scopes = config.get("scopes", ["https://www.googleapis.com/auth/gmail.send"])

        self.creds = None
        self._authenticate(logger, Request, Credentials, InstalledAppFlow)

    def _authenticate(self, logger, Request, Credentials, InstalledAppFlow) -> None:
        """Authenticate with Gmail OAuth2."""
        # Load existing token if available
        if os.path.exists(self.token_file):
            self.creds = Credentials.from_authorized_user_file(self.token_file, self.scopes)

        # If no valid credentials, authenticate
        if not self.creds or not self.creds.valid:
            if self.creds and self.creds.expired and self.creds.refresh_token:
                logger.info("Refreshing Gmail OAuth2 token...")
                self.creds.refresh(Request())
            else:
                if not os.path.exists(self.credentials_file):
                    raise FileNotFoundError(
                        f"Gmail credentials file not found: {self.credentials_file}\n"
                        "Please download OAuth2 credentials from Google Cloud Console."
                    )
                logger.info("Starting Gmail OAuth2 authorization flow...")
                logger.info("A browser window will open for authorization.")
                flow = InstalledAppFlow.from_client_secrets_file(
                    self.credentials_file, self.scopes
                )
                self.creds = flow.run_local_server(port=0)

            # Save the token for future use
            with open(self.token_file, "w") as token:
                token.write(self.creds.to_json())
            logger.info(f"Gmail OAuth2 token saved to: {self.token_file}")

    def send(
        self,
        from_addr: str,
        to_addr: str,
        subject: str,
        body: str,
        logger: logging.Logger
    ) -> bool:
        """Send email using Gmail API."""
        try:
            from googleapiclient.discovery import build
            from googleapiclient.errors import HttpError
        except ImportError:
            raise ImportError(
                "Gmail API requires google-api-python-client. "
                "Install with: pip install google-api-python-client"
            )

        try:
            # Build the Gmail service
            service = build("gmail", "v1", credentials=self.creds)

            # Create the email message
            msg = MIMEMultipart()
            msg["From"] = from_addr
            msg["To"] = to_addr
            msg["Subject"] = subject
            msg.attach(MIMEText(body, "plain", "utf-8"))

            # Encode the message
            raw_message = base64.urlsafe_b64encode(msg.as_bytes()).decode("utf-8")

            # Send the email
            service.users().messages().send(
                userId="me",
                body={"raw": raw_message}
            ).execute()

            logger.info(f"Email sent successfully to: {to_addr}")
            return True

        except HttpError as e:
            logger.error(f"Failed to send email to {to_addr}: {e}")
            return False
        except Exception as e:
            logger.error(f"Failed to send email to {to_addr}: {e}")
            return False

    def close(self) -> None:
        """No cleanup needed for Gmail OAuth2."""
        pass


# =============================================================================
# Outlook OAuth2 Email Sender
# =============================================================================

class OutlookOAuth2Sender(EmailSender):
    """Email sender using Microsoft OAuth2 authentication."""

    def __init__(self, config: dict, logger: logging.Logger):
        """
        Initialize Outlook OAuth2 sender.

        Args:
            config: Outlook OAuth2 configuration dictionary
            logger: Logger instance
        """
        try:
            import msal
        except ImportError:
            raise ImportError(
                "Outlook OAuth2 requires msal. "
                "Install with: pip install msal"
            )

        self.client_id = config["client_id"]
        self.client_secret = config.get("client_secret")
        self.tenant_id = config.get("tenant_id", "common")
        self.token_cache_file = config.get("token_cache_file", "outlook_token_cache.json")
        self.scopes = config.get("scopes", ["https://graph.microsoft.com/Mail.Send"])

        self.access_token = None
        self._authenticate(logger, msal)

    def _load_token_cache(self, msal) -> Any:
        """Load token cache from file."""
        cache = msal.SerializableTokenCache()
        if os.path.exists(self.token_cache_file):
            with open(self.token_cache_file, "r") as f:
                cache.deserialize(f.read())
        return cache

    def _save_token_cache(self, cache: Any) -> None:
        """Save token cache to file."""
        if cache.has_state_changed:
            with open(self.token_cache_file, "w") as f:
                f.write(cache.serialize())

    def _authenticate(self, logger, msal) -> None:
        """Authenticate with Microsoft OAuth2."""
        cache = self._load_token_cache(msal)

        authority = f"https://login.microsoftonline.com/{self.tenant_id}"

        # Create MSAL application
        app = msal.PublicClientApplication(
            self.client_id,
            authority=authority,
            token_cache=cache
        )

        # Try to get token from cache
        accounts = app.get_accounts()
        if accounts:
            logger.info("Found cached Outlook account, attempting silent authentication...")
            result = app.acquire_token_silent(self.scopes, account=accounts[0])
            if result and "access_token" in result:
                self.access_token = result["access_token"]
                self._save_token_cache(cache)
                logger.info("Outlook OAuth2 authentication successful (from cache)")
                return

        # Interactive authentication
        logger.info("Starting Outlook OAuth2 authorization flow...")
        logger.info("A browser window will open for authorization.")

        result = app.acquire_token_interactive(
            scopes=self.scopes,
            prompt="select_account"
        )

        if "access_token" in result:
            self.access_token = result["access_token"]
            self._save_token_cache(cache)
            logger.info(f"Outlook OAuth2 token cached to: {self.token_cache_file}")
        else:
            error = result.get("error_description", result.get("error", "Unknown error"))
            raise RuntimeError(f"Outlook OAuth2 authentication failed: {error}")

    def send(
        self,
        from_addr: str,
        to_addr: str,
        subject: str,
        body: str,
        logger: logging.Logger
    ) -> bool:
        """Send email using Microsoft Graph API."""
        try:
            import requests
        except ImportError:
            raise ImportError(
                "Outlook OAuth2 requires requests. "
                "Install with: pip install requests"
            )

        try:
            # Microsoft Graph API endpoint for sending email
            endpoint = "https://graph.microsoft.com/v1.0/me/sendMail"

            # Prepare email payload
            email_payload = {
                "message": {
                    "subject": subject,
                    "body": {
                        "contentType": "Text",
                        "content": body
                    },
                    "toRecipients": [
                        {
                            "emailAddress": {
                                "address": to_addr
                            }
                        }
                    ]
                },
                "saveToSentItems": "true"
            }

            # Send the request
            headers = {
                "Authorization": f"Bearer {self.access_token}",
                "Content-Type": "application/json"
            }

            response = requests.post(endpoint, headers=headers, json=email_payload)

            if response.status_code == 202:
                logger.info(f"Email sent successfully to: {to_addr}")
                return True
            else:
                logger.error(
                    f"Failed to send email to {to_addr}: "
                    f"Status {response.status_code}, {response.text}"
                )
                return False

        except Exception as e:
            logger.error(f"Failed to send email to {to_addr}: {e}")
            return False

    def close(self) -> None:
        """No cleanup needed for Outlook OAuth2."""
        pass


# =============================================================================
# Dry Run Email Sender (for testing)
# =============================================================================

class DryRunSender(EmailSender):
    """Email sender that only logs without sending."""

    def send(
        self,
        from_addr: str,
        to_addr: str,
        subject: str,
        body: str,
        logger: logging.Logger
    ) -> bool:
        """Log email details without sending."""
        logger.info(f"[DRY RUN] Would send email to: {to_addr}")
        logger.debug(f"[DRY RUN] Subject: {subject}")
        logger.debug(f"[DRY RUN] Body preview: {body[:100]}...")
        return True

    def close(self) -> None:
        """No cleanup needed for dry run."""
        pass


# =============================================================================
# Email Sender Factory
# =============================================================================

def create_email_sender(
    config: dict,
    logger: logging.Logger,
    dry_run: bool = False
) -> EmailSender:
    """
    Create an email sender based on configuration.

    Args:
        config: Configuration dictionary
        logger: Logger instance
        dry_run: If True, return a DryRunSender

    Returns:
        EmailSender instance
    """
    if dry_run:
        return DryRunSender()

    auth_type = config.get("auth_type", "basic").lower()

    if auth_type == "basic":
        logger.info("Using Basic SMTP authentication")
        return BasicSMTPSender(config["smtp"])

    elif auth_type == "gmail_oauth2":
        logger.info("Using Gmail OAuth2 authentication")
        return GmailOAuth2Sender(config["gmail_oauth2"], logger)

    elif auth_type == "outlook_oauth2":
        logger.info("Using Outlook OAuth2 authentication")
        return OutlookOAuth2Sender(config["outlook_oauth2"], logger)

    else:
        raise ValueError(f"Unknown auth_type: {auth_type}")


# =============================================================================
# Utility Functions
# =============================================================================

def setup_logging(config: dict) -> logging.Logger:
    """
    Setup logging based on configuration.

    Args:
        config: Configuration dictionary containing logging settings

    Returns:
        Configured logger instance
    """
    log_config = config.get("logging", {})
    log_level = getattr(logging, log_config.get("level", "INFO").upper(), logging.INFO)
    log_file = log_config.get("file", "")

    logger = logging.getLogger("email_scheduler")
    logger.setLevel(log_level)

    # Clear existing handlers
    logger.handlers.clear()

    # Console handler
    console_handler = logging.StreamHandler()
    console_handler.setLevel(log_level)
    console_format = logging.Formatter("[%(asctime)s] %(levelname)s: %(message)s")
    console_handler.setFormatter(console_format)
    logger.addHandler(console_handler)

    # File handler (optional)
    if log_file:
        file_handler = logging.FileHandler(log_file)
        file_handler.setLevel(log_level)
        file_format = logging.Formatter("[%(asctime)s] %(levelname)s: %(message)s")
        file_handler.setFormatter(file_format)
        logger.addHandler(file_handler)

    return logger


def load_config(config_path: str) -> dict:
    """
    Load configuration from YAML file.

    Args:
        config_path: Path to the configuration file

    Returns:
        Configuration dictionary

    Raises:
        FileNotFoundError: If config file doesn't exist
        yaml.YAMLError: If config file is invalid
    """
    with open(config_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def load_processed_files(record_path: str) -> dict:
    """
    Load the record of processed files.

    Args:
        record_path: Path to the processed files record JSON

    Returns:
        Dictionary with file paths as keys and processing info as values
    """
    if os.path.exists(record_path):
        with open(record_path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_processed_files(record_path: str, processed: dict) -> None:
    """
    Save the record of processed files.

    Args:
        record_path: Path to the processed files record JSON
        processed: Dictionary with file paths as keys and processing info as values
    """
    with open(record_path, "w", encoding="utf-8") as f:
        json.dump(processed, f, indent=2, ensure_ascii=False)


def get_file_signature(file_path: Path) -> str:
    """
    Get a signature for a file based on its modification time and size.

    Args:
        file_path: Path to the file

    Returns:
        String signature combining mtime and size
    """
    stat = file_path.stat()
    return f"{stat.st_mtime}_{stat.st_size}"


def find_xlsx_files(workspace_path: str) -> list[Path]:
    """
    Find all xlsx files in the workspace directory.

    Args:
        workspace_path: Path to the workspace directory

    Returns:
        List of Path objects for xlsx files
    """
    workspace = Path(workspace_path)
    if not workspace.exists():
        raise FileNotFoundError(f"Workspace directory not found: {workspace_path}")

    xlsx_files = list(workspace.glob("*.xlsx"))
    # Filter out temporary files (starting with ~$)
    xlsx_files = [f for f in xlsx_files if not f.name.startswith("~$")]
    return xlsx_files


def fill_template(template: str, row_data: dict[str, Any]) -> str:
    """
    Fill a template string with values from row data.

    Args:
        template: Template string containing {placeholder} patterns
        row_data: Dictionary mapping column names to values

    Returns:
        Filled template string
    """
    result = template
    for key, value in row_data.items():
        placeholder = "{" + key + "}"
        # Convert value to string, handle None values
        str_value = str(value) if value is not None else ""
        result = result.replace(placeholder, str_value)
    return result


def find_leader_columns(headers: list[str]) -> list[str]:
    """
    Find all column names containing 'Leader' (case-insensitive).

    Args:
        headers: List of column header names

    Returns:
        List of column names containing 'Leader'
    """
    return [h for h in headers if h and "leader" in h.lower()]


def is_valid_email(email: str) -> bool:
    """
    Check if a string is a valid email address.

    Args:
        email: String to validate

    Returns:
        True if valid email format, False otherwise
    """
    if not email or not isinstance(email, str):
        return False
    # Simple email validation regex
    pattern = r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$"
    return bool(re.match(pattern, email.strip()))


def read_xlsx_data(file_path: Path, logger: logging.Logger) -> list[dict[str, Any]]:
    """
    Read data from an xlsx file.

    Args:
        file_path: Path to the xlsx file
        logger: Logger instance

    Returns:
        List of dictionaries, each representing a row with column names as keys
    """
    logger.info(f"Reading file: {file_path}")

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        logger.warning(f"File {file_path} is empty")
        return []

    # First row is headers
    headers = [str(h).strip() if h else f"Column_{i}" for i, h in enumerate(rows[0])]
    logger.debug(f"Headers: {headers}")

    # Convert remaining rows to dictionaries
    data = []
    for row_idx, row in enumerate(rows[1:], start=2):
        row_dict = {}
        for col_idx, value in enumerate(row):
            if col_idx < len(headers):
                row_dict[headers[col_idx]] = value
        data.append({"row_number": row_idx, "data": row_dict})

    logger.info(f"Read {len(data)} rows from {file_path}")
    wb.close()
    return data


# =============================================================================
# Main Processing Functions
# =============================================================================

def process_xlsx_file(
    file_path: Path,
    config: dict,
    sender: EmailSender,
    logger: logging.Logger
) -> dict[str, int]:
    """
    Process a single xlsx file and send emails.

    Args:
        file_path: Path to the xlsx file
        config: Configuration dictionary
        sender: EmailSender instance
        logger: Logger instance

    Returns:
        Dictionary with statistics (sent, failed, skipped)
    """
    stats = {"sent": 0, "failed": 0, "skipped": 0}

    # Read xlsx data
    try:
        rows = read_xlsx_data(file_path, logger)
    except Exception as e:
        logger.error(f"Failed to read file {file_path}: {e}")
        return stats

    if not rows:
        return stats

    # Get email configuration
    email_config = config["email"]
    subject_template = email_config["subject"]
    body_template = email_config["body"]
    from_addr = email_config["from_address"]

    # Get headers from first row's data
    if rows:
        headers = list(rows[0]["data"].keys())
        leader_columns = find_leader_columns(headers)

        if not leader_columns:
            logger.warning(f"No 'Leader' columns found in {file_path}")
            return stats

        logger.info(f"Found Leader columns: {leader_columns}")

    # Process each row
    for row_info in rows:
        row_num = row_info["row_number"]
        row_data = row_info["data"]

        # Fill templates with row data
        subject = fill_template(subject_template, row_data)
        body = fill_template(body_template, row_data)

        # Send to all Leader email addresses in this row
        for leader_col in leader_columns:
            email_addr = row_data.get(leader_col)

            if not is_valid_email(str(email_addr) if email_addr else ""):
                logger.debug(
                    f"Row {row_num}: Skipping invalid email in column '{leader_col}': {email_addr}"
                )
                stats["skipped"] += 1
                continue

            email_addr = str(email_addr).strip()

            logger.info(f"Row {row_num}: Sending email to {email_addr} (column: {leader_col})")

            if sender.send(from_addr, email_addr, subject, body, logger):
                stats["sent"] += 1
            else:
                stats["failed"] += 1

    return stats


def find_new_xlsx_files(
    workspace_path: str,
    processed_record: dict,
    logger: logging.Logger
) -> list[Path]:
    """
    Find xlsx files that are new or modified since last processing.

    Args:
        workspace_path: Path to the workspace directory
        processed_record: Dictionary of previously processed files
        logger: Logger instance

    Returns:
        List of new/modified xlsx file paths
    """
    all_files = find_xlsx_files(workspace_path)
    new_files = []

    for file_path in all_files:
        file_key = str(file_path.absolute())
        current_signature = get_file_signature(file_path)

        if file_key not in processed_record:
            logger.debug(f"New file detected: {file_path.name}")
            new_files.append(file_path)
        elif processed_record[file_key].get("signature") != current_signature:
            logger.debug(f"Modified file detected: {file_path.name}")
            new_files.append(file_path)

    return new_files


def process_new_files(
    config: dict,
    sender: EmailSender,
    logger: logging.Logger,
    dry_run: bool = False
) -> dict[str, int]:
    """
    Check for and process new xlsx files.

    Args:
        config: Configuration dictionary
        sender: EmailSender instance
        logger: Logger instance
        dry_run: Whether in dry run mode

    Returns:
        Dictionary with statistics (sent, failed, skipped, files_processed)
    """
    workspace_path = config.get("workspace", {}).get("path", "./workspace")
    scheduler_config = config.get("scheduler", {})
    record_path = scheduler_config.get("processed_files_record", "./processed_files.json")

    # Load processed files record
    processed_record = load_processed_files(record_path)

    # Find new files
    try:
        new_files = find_new_xlsx_files(workspace_path, processed_record, logger)
    except FileNotFoundError as e:
        logger.error(str(e))
        return {"sent": 0, "failed": 0, "skipped": 0, "files_processed": 0}

    if not new_files:
        logger.debug("No new xlsx files found")
        return {"sent": 0, "failed": 0, "skipped": 0, "files_processed": 0}

    logger.info(f"Found {len(new_files)} new/modified xlsx file(s) to process")

    total_stats = {"sent": 0, "failed": 0, "skipped": 0, "files_processed": 0}

    for xlsx_file in new_files:
        logger.info(f"Processing: {xlsx_file.name}")
        stats = process_xlsx_file(xlsx_file, config, sender, logger)

        for key in ["sent", "failed", "skipped"]:
            total_stats[key] += stats[key]

        # Mark file as processed
        file_key = str(xlsx_file.absolute())
        processed_record[file_key] = {
            "signature": get_file_signature(xlsx_file),
            "processed_at": time.strftime("%Y-%m-%d %H:%M:%S"),
            "stats": stats
        }
        total_stats["files_processed"] += 1

        # Save after each file in case of interruption
        if not dry_run:
            save_processed_files(record_path, processed_record)

    return total_stats


# Global flag for graceful shutdown
_shutdown_requested = False


def _signal_handler(signum, frame):
    """Handle shutdown signals."""
    global _shutdown_requested
    _shutdown_requested = True


def run_scheduler(config: dict, logger: logging.Logger, dry_run: bool = False) -> None:
    """
    Run the scheduler loop to periodically check for new xlsx files.

    Args:
        config: Configuration dictionary
        logger: Logger instance
        dry_run: Whether in dry run mode
    """
    global _shutdown_requested

    scheduler_config = config.get("scheduler", {})
    check_interval = scheduler_config.get("check_interval", 86400)  # Default: 1 day

    logger.info("=" * 50)
    logger.info("Starting Email Scheduler in watch mode")
    logger.info(f"Check interval: {check_interval} seconds ({check_interval / 3600:.1f} hours)")
    logger.info("Press Ctrl+C to stop")
    logger.info("=" * 50)

    # Setup signal handlers for graceful shutdown
    signal.signal(signal.SIGINT, _signal_handler)
    signal.signal(signal.SIGTERM, _signal_handler)

    # Create email sender
    try:
        sender = create_email_sender(config, logger, dry_run)
    except Exception as e:
        logger.error(f"Failed to initialize email sender: {e}")
        return

    try:
        while not _shutdown_requested:
            logger.info("Checking for new xlsx files...")

            stats = process_new_files(config, sender, logger, dry_run)

            if stats["files_processed"] > 0:
                logger.info(f"Processed {stats['files_processed']} file(s): "
                           f"sent={stats['sent']}, failed={stats['failed']}, skipped={stats['skipped']}")
            else:
                logger.info("No new files to process")

            # Wait for next check interval
            logger.info(f"Next check in {check_interval} seconds...")

            # Sleep in small intervals to allow for graceful shutdown
            sleep_interval = min(check_interval, 10)
            elapsed = 0
            while elapsed < check_interval and not _shutdown_requested:
                time.sleep(sleep_interval)
                elapsed += sleep_interval

    finally:
        sender.close()
        logger.info("Scheduler stopped")


def main():
    """Main entry point for the email scheduler."""
    # Parse command line arguments
    parser = argparse.ArgumentParser(
        description="Send emails based on xlsx data and configurable templates"
    )
    parser.add_argument(
        "--config",
        default="config.yaml",
        help="Path to configuration file (default: config.yaml)"
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Preview emails without sending"
    )
    parser.add_argument(
        "--watch",
        action="store_true",
        help="Run in watch mode: periodically check for new xlsx files"
    )
    args = parser.parse_args()

    # Load configuration
    try:
        config = load_config(args.config)
    except FileNotFoundError:
        print(f"Error: Configuration file not found: {args.config}")
        sys.exit(1)
    except yaml.YAMLError as e:
        print(f"Error: Invalid configuration file: {e}")
        sys.exit(1)

    # Setup logging
    logger = setup_logging(config)

    # Determine dry run mode (CLI argument takes precedence)
    dry_run = args.dry_run or config.get("dry_run", False)

    if dry_run:
        logger.info("Running in DRY RUN mode - no emails will be sent")

    # Check if watch mode is requested (CLI argument or config)
    scheduler_config = config.get("scheduler", {})
    watch_mode = args.watch or scheduler_config.get("enabled", False)

    if watch_mode:
        run_scheduler(config, logger, dry_run)
    else:
        # Original one-time processing mode
        # Create email sender
        try:
            sender = create_email_sender(config, logger, dry_run)
        except Exception as e:
            logger.error(f"Failed to initialize email sender: {e}")
            sys.exit(1)

        # Find xlsx files
        workspace_path = config.get("workspace", {}).get("path", "./workspace")

        try:
            xlsx_files = find_xlsx_files(workspace_path)
        except FileNotFoundError as e:
            logger.error(str(e))
            sys.exit(1)

        if not xlsx_files:
            logger.warning(f"No xlsx files found in {workspace_path}")
            sys.exit(0)

        logger.info(f"Found {len(xlsx_files)} xlsx file(s) to process")

        # Process each xlsx file
        total_stats = {"sent": 0, "failed": 0, "skipped": 0}

        try:
            for xlsx_file in xlsx_files:
                logger.info(f"Processing: {xlsx_file.name}")
                stats = process_xlsx_file(xlsx_file, config, sender, logger)

                for key in total_stats:
                    total_stats[key] += stats[key]
        finally:
            # Clean up sender resources
            sender.close()

        # Print summary
        logger.info("=" * 50)
        logger.info("Processing Complete!")
        logger.info(f"Total emails sent: {total_stats['sent']}")
        logger.info(f"Total emails failed: {total_stats['failed']}")
        logger.info(f"Total emails skipped: {total_stats['skipped']}")

        if total_stats["failed"] > 0:
            sys.exit(1)


if __name__ == "__main__":
    main()
